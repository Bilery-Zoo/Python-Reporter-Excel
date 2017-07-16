# !/usr/bin/python
# -*- coding: utf-8 -*-


"""
create_author : Bilery Zoo(652645572@qq.com)
create_time   : 2017-04-21
program       : *_* Write Excel in a Report Format *_*
"""


import re
import openpyxl


class ExcelHandler:
    """
    Handler of general demand to write Excel file.
    """
    def __init__(self, writer="writer.xlsx"):
        """
        Set an Excel file.
        :param writer: string, default : an Excel file named "writer.xlsx" at CWD
            full directory of the Excel file to write out.
        """
        self.wkbook = openpyxl.Workbook()
        self.writer = writer

    def __repr__(self):
        printer = 'o(>﹏<)o ......Excel old driver write {0} out...... o(^o^)o'.format(self.writer)
        return printer

    def createsheet(self, sheetname):
        """
        Create an Excel work sheet.
        :param sheetname: string
            Excel work sheet name.
        :return: sheet object.
        """
        try:
            del self.wkbook["Sheet"]
        except:
            pass
        wsheet = self.wkbook.create_sheet(sheetname)
        return wsheet

    def definetitle(self, wsheet, title, cellspace, heighter=None,
                    name='Calibri', size=12, bold=True, italic=False, underline='none', color="0070C0",
                    fill_type=None, fgColor="FFFFFF",
                    horizontal='center', vertical='center'):
        """
        Define an Excel sheet title.
        :param wsheet: sheet object
            Function createsheet() returned.
        :param title: int, string and other formats Excel support
            Excel sheet title.
        :param cellspace: string
            Excel cell intervals the "title" to write seporated by ':'(eg : "A1 : C11").
        :param heighter: int, default : none(simply write)
            Excel row height.
        :param name: see " http://openpyxl.readthedocs.io/en/default/styles.html ".
        :param size: see " http://openpyxl.readthedocs.io/en/default/styles.html ".
        :param bold: see " http://openpyxl.readthedocs.io/en/default/styles.html ".
        :param italic: see " http://openpyxl.readthedocs.io/en/default/styles.html ".
        :param underline: see " http://openpyxl.readthedocs.io/en/default/styles.html ".
        :param color: see " http://openpyxl.readthedocs.io/en/default/styles.html ".
        :param fill_type: see " http://openpyxl.readthedocs.io/en/default/styles.html ".
        :param fgColor: see " http://openpyxl.readthedocs.io/en/default/styles.html ".
        :param horizontal: see " http://openpyxl.readthedocs.io/en/default/styles.html ".
        :param vertical: see " http://openpyxl.readthedocs.io/en/default/styles.html ".
        :return: none.
        """
        font = openpyxl.styles.Font(name=name, size=size, bold=bold, italic=italic, underline=underline, color=color)
        fill = openpyxl.styles.PatternFill(fill_type=fill_type, fgColor=fgColor)
        alig = openpyxl.styles.Alignment(horizontal=horizontal, vertical=vertical)
        cell = cellspace.split(':')[0]
        wsheet.merge_cells(cellspace)
        wsheet[cell] = title
        wsheet[cell].font = font
        wsheet[cell].alignment = alig
        if heighter:
            row = re.findall(r"\d+", cellspace)
            wsheet.row_dimensions[int(row[0])].height = heighter
            wsheet.row_dimensions[int(row[1])].height = heighter

    def setheader(self, wsheet, header, beginer, heighter=None, wider=None,
                  name='Calibri', size=12, bold=False, italic=False, underline='none', color="FF000000",
                  fill_type="solid", fgColor="FFFF00",
                  horizontal='center', vertical='center'):
        """
        Set an Excel sheet title(column names).
        :param wsheet: sheet object
            Function createsheet() returned.
        :param header: list
            Excel sheet title(column names).
        :param beginer: list
            Position([row number, column number]) of the sheet title(column names) starts to write.
        :param heighter: int, default : none(simply write)
            Excel row height.
        :param wider: list, default : none(auto fit by cell data length)
            Excel column width.
        :param name: see " http://openpyxl.readthedocs.io/en/default/styles.html ".
        :param size: see " http://openpyxl.readthedocs.io/en/default/styles.html ".
        :param bold: see " http://openpyxl.readthedocs.io/en/default/styles.html ".
        :param italic: see " http://openpyxl.readthedocs.io/en/default/styles.html ".
        :param underline: see " http://openpyxl.readthedocs.io/en/default/styles.html ".
        :param color: see " http://openpyxl.readthedocs.io/en/default/styles.html ".
        :param fill_type: see " http://openpyxl.readthedocs.io/en/default/styles.html ".
        :param fgColor: see " http://openpyxl.readthedocs.io/en/default/styles.html ".
        :param horizontal: see " http://openpyxl.readthedocs.io/en/default/styles.html ".
        :param vertical: see " http://openpyxl.readthedocs.io/en/default/styles.html ".
        :return: none.
        """
        font = openpyxl.styles.Font(name=name, size=size, bold=bold, italic=italic, underline=underline, color=color)
        fill = openpyxl.styles.PatternFill(fill_type=fill_type, fgColor=fgColor)
        alig = openpyxl.styles.Alignment(horizontal=horizontal, vertical=vertical)
        rownber = beginer[0]
        colnber = beginer[1]
        counter = len(header)
        stor = list(zip(range(colnber, counter + 2), header))
        for i, j in stor:
            wsheet.cell(row=rownber, column=i).value = j
            wsheet.cell(row=rownber, column=i).font = font
            wsheet.cell(row=rownber, column=i).fill = fill
            wsheet.cell(row=rownber, column=i).alignment = alig
        if heighter:
            wsheet.row_dimensions[rownber].height = heighter
        if wider:
            colitem = list(zip(range(len(header)), wider))
            for col, wid in colitem:
                colth = openpyxl.utils.get_column_letter(col + colnber)
                wsheet.column_dimensions[colth].width = wid
        elif not wider:
            for i in range(len(header)):
                col = openpyxl.utils.get_column_letter(i + colnber)
                wid = len(str(header[i])) + 5
                wsheet.column_dimensions[col].width = wid

    def insertrecord(self, wsheet, record):
        """
        Insert record into an Excel sheet row by row.
        :param wsheet: sheet object
            Function createsheet() returned.
        :param record: list
            Excel data records, [[...], [...], ...].
        :return: none.
        """
        for row in record:
            wsheet.append(row)

    def mergecolcellwrite(self, wsheet, record, beginer, mercols=None, isautoFitwid=False):
        """
        Optimized function of writing Excel record: merge same data cells by column, auto fit column width.
        :param wsheet: sheet object
            Function createsheet() returned.
        :param record: list
            Excel data records, [[...], [...], ...].
        :param beginer: list
            Position([row number, column number]) of the sheet "record" starts to write.
        :param mercols: list, default : none(simply write)
            Columns to merge same data cells when write.
        :param isautoFitwid: bool, default : False(simply write)
            Whether to auto fit column width by cell data length.
        :return: none.
        """
        rownber = beginer[0]
        colnber = beginer[1]
        if isautoFitwid:
            for i in range(len(record[0])):
                col = openpyxl.utils.get_column_letter(colnber + i)
                sto = []
                for j in record:
                    sto.append(len(str(j[i])))
                wid = max(sto) + 3
                wsheet.column_dimensions[col].width = wid
        for colindex in range(len(record[0])):
            store = []
            for rowindex in range(len(record)):
                value = record[rowindex][colindex]
                if not mercols:
                    wsheet.cell(row=rownber, column=colnber, value=value)
                elif mercols:
                    for mergcol in mercols:
                        if mergcol != colnber:
                            wsheet.cell(row=rownber, column=colnber, value=value)
                        elif mergcol == colnber:
                            try:
                                if rowindex == 0:
                                    store.append([rownber, value])
                                elif rowindex == len(record) - 1:
                                    if store[-1][1] != value:
                                        precell = openpyxl.utils.get_column_letter(colnber) + str(store[0][0])
                                        sufcell = openpyxl.utils.get_column_letter(colnber) + str(store[-1][0])
                                        wsheet.merge_cells("{0}:{1}".format(precell, sufcell))
                                        wsheet.cell(row=store[0][0], column=colnber, value=store[0][1])
                                        wsheet.cell(row=rownber, column=colnber, value=value)
                                    elif store[-1][1] == value:
                                        store.append([rownber, value])
                                        precell = openpyxl.utils.get_column_letter(colnber) + str(store[0][0])
                                        sufcell = openpyxl.utils.get_column_letter(colnber) + str(store[-1][0])
                                        wsheet.merge_cells("{0}:{1}".format(precell, sufcell))
                                        wsheet.cell(row=store[0][0], column=colnber, value=value)
                                elif store[-1][1] == value:
                                    store.append([rownber, value])
                                elif store[-1][1] != value:
                                    try:
                                        precell = openpyxl.utils.get_column_letter(colnber) + str(store[0][0])
                                        sufcell = openpyxl.utils.get_column_letter(colnber) + str(store[-1][0])
                                        wsheet.merge_cells("{0}:{1}".format(precell, sufcell))
                                        wsheet.cell(row=store[0][0], column=colnber, value=store[0][1])
                                    except:
                                        pass
                                    finally:
                                        store = []
                                        store.append([rownber, value])
                            except:
                                pass
                rownber += 1
            colnber += 1
            rownber = beginer[0]

    def savexcel(self):
        """
        Save out Excel file.
        :return: none.
        """
        self.wkbook.save(self.writer)


# self test
if __name__ == "__main__":
    MyEH = ExcelHandler("/home/student/wb.xlsx")

    title = "Hello World\nHello Handler"
    header = ['name', 'rank', 'logo']
    record = [["Linux", 0, "penguin"],
              ["MySQL", 1, "dolphin"],
              ["MySQL", 1, "hawk"],
              ["Python", 2, "python"],
              ["Python", 2, "mouse"],
              ["Python", 2, "rabbit"]]

    wsheet_l = MyEH.createsheet("Hello World")
    MyEH.definetitle(wsheet_l, title, "A32:C33", heighter=23)
    MyEH.setheader(wsheet_l, header, [36, 1], heighter=23, wider=[11, 23, 11])
    MyEH.mergecolcellwrite(wsheet_l, record, [37, 1], [1, 2])

    wsheet_r = MyEH.createsheet("Hello Handler")
    MyEH.definetitle(wsheet_r, title, "A32:C33", heighter=23)
    MyEH.setheader(wsheet_r, header, [36, 1], heighter=23, wider=[11, 23, 11])
    MyEH.insertrecord(wsheet_r, record)

    MyEH.savexcel()
